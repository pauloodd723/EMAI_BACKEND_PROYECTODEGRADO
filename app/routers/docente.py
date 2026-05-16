import json
from datetime import date
from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks, Request
from sqlalchemy.orm import Session
from sqlalchemy import func
from fastapi.responses import Response
from ..database import get_db
from ..models import (
    User, Course, Student, Exam, ExamResult,
    SupportMessage, UserRoleEnum, ExamStatusEnum,
)
from ..security import require_role, get_current_user
from ..services.ocr_service import (
    extract_text_from_images,
    detect_math_problems,
    calculate_final_score,
    generate_teaching_plan,
)

router = APIRouter(prefix="/docente", tags=["docente"])
require_docente = require_role(UserRoleEnum.docente)

PASSING_GRADE = 3.0


def _teacher_institution(current_user: User) -> str:
    if not current_user.institution_id:
        raise HTTPException(status_code=400, detail="Docente sin institución")
    return current_user.institution_id


def _course_to_dict(c: Course) -> dict:
    return {
        "id": c.id, "name": c.name, "grade": c.grade, "group": c.group,
        "institutionId": c.institution_id, "teacherId": c.teacher_id,
        "createdAt": c.created_at.isoformat(),
    }


def _student_to_dict(s: Student) -> dict:
    return {
        "id": s.id, "fullName": s.full_name, "photoUrl": s.photo_url,
        "courseId": s.course_id, "institutionId": s.institution_id,
        "createdBy": s.created_by_id,
        "createdAt": s.created_at.isoformat(),
        "updatedAt": s.updated_at.isoformat(),
    }


def _result_to_dict(result: ExamResult, exam_name: str = "") -> dict:
    image_urls = json.loads(result.image_urls or "[]")
    problems_raw = json.loads(result.problems_json or "[]")

    problems = None
    if problems_raw:
        problems = [
            {
                "id": p.get("id"),
                "section": p.get("section", ""),
                "problemText": p.get("problemText") or p.get("problem_text", ""),
                "originalText": p.get("originalText") or p.get("original_text") or p.get("operation", ""),
                "operation": p.get("operation", ""),
                "type": p.get("type", "complex_math"),
                "subType": p.get("subType") or p.get("sub_type", "+"),
                "studentAnswer": p.get("studentAnswer") or p.get("student_answer", "-"),
                "correctAnswer": p.get("correctAnswer") or p.get("correct_answer", ""),
                "isCorrect": p.get("isCorrect") if p.get("isCorrect") is not None else p.get("is_correct"),
                "teacherOverride": p.get("teacherOverride", False),
            }
            for p in problems_raw
        ]

    # Separar notas del docente del plan pedagógico
    raw_notes = result.teacher_notes or ""
    teacher_notes_only = ""
    teaching_plan_text = ""
    if "---PLAN---" in raw_notes:
        parts = raw_notes.split("---PLAN---", 1)
        teacher_notes_only = parts[0].strip()
        teaching_plan_text = parts[1].strip()
    else:
        teacher_notes_only = raw_notes

    return {
        "id": result.id,
        "examId": result.exam_id,
        "examName": exam_name,
        "studentId": result.student_id,
        "imageUrls": image_urls,
        "ocrRawText": result.ocr_raw_text,
        "problems": problems,
        "finalScore": result.final_score,
        "gradeColor": result.grade_color,
        "teacherNotes": teacher_notes_only,
        "teachingPlan": teaching_plan_text,
        "status": result.status.value if result.status else "pending",
        "createdAt": result.created_at.isoformat(),
        "updatedAt": result.updated_at.isoformat(),
    }


# ─── STATS ────────────────────────────────────────────────────────────────────
@router.get("/stats")
def get_stats(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_docente),
):
    inst_id = _teacher_institution(current_user)
    today = date.today()

    courses = db.query(Course).filter(
        Course.institution_id == inst_id,
        Course.teacher_id == current_user.id,
    ).all()
    course_ids = [c.id for c in courses]

    total_students = (
        db.query(Student).filter(Student.course_id.in_(course_ids)).count()
        if course_ids else 0
    )
    exam_ids_q = db.query(Exam.id).filter(Exam.teacher_id == current_user.id)
    total_exams = db.query(ExamResult).filter(ExamResult.exam_id.in_(exam_ids_q)).count()
    exams_today = db.query(ExamResult).filter(
        ExamResult.exam_id.in_(exam_ids_q),
        func.date(ExamResult.created_at) == today,
    ).count()
    scores_q = db.query(ExamResult.final_score).filter(
        ExamResult.final_score.isnot(None),
        ExamResult.exam_id.in_(exam_ids_q),
    ).all()
    scores = [r.final_score for r in scores_q if r.final_score]
    promedio = round(sum(scores) / len(scores), 2) if scores else None

    return {
        "totalEstudiantes": total_students,
        "totalExamenes": total_exams,
        "examenesHoy": exams_today,
        "promedioGeneral": promedio,
    }


# ─── CURSOS ───────────────────────────────────────────────────────────────────
@router.get("/cursos")
def list_my_courses(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_docente),
):
    inst_id = _teacher_institution(current_user)
    courses = db.query(Course).filter(
        Course.institution_id == inst_id,
        Course.teacher_id == current_user.id,
    ).order_by(Course.grade, Course.group).all()
    return [_course_to_dict(c) for c in courses]


@router.get("/cursos/todos")
def list_all_courses(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_docente),
):
    inst_id = _teacher_institution(current_user)
    courses = db.query(Course).filter(
        Course.institution_id == inst_id,
    ).order_by(Course.grade, Course.group).all()
    return [_course_to_dict(c) for c in courses]


# ─── ESTUDIANTES ──────────────────────────────────────────────────────────────
@router.get("/cursos/{course_id}/estudiantes")
def list_students(
    course_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_docente),
):
    students = db.query(Student).filter(
        Student.course_id == course_id,
        Student.institution_id == current_user.institution_id,
    ).order_by(Student.full_name).all()
    return [_student_to_dict(s) for s in students]


@router.get("/estudiantes/{student_id}")
def get_student(
    student_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_docente),
):
    s = db.query(Student).filter(
        Student.id == student_id,
        Student.institution_id == current_user.institution_id,
    ).first()
    if not s:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    return _student_to_dict(s)


@router.post("/estudiantes", status_code=201)
async def create_student(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_docente),
):
    inst_id = _teacher_institution(current_user)
    body = await request.json()
    full_name = (body.get("fullName") or body.get("full_name") or "").strip()
    course_id = body.get("courseId") or body.get("course_id")
    photo_url = body.get("photoUrl") or body.get("photo_url")

    if not full_name:
        raise HTTPException(status_code=400, detail="El nombre es requerido")

    student = Student(
        full_name=full_name, course_id=course_id,
        institution_id=inst_id, created_by_id=current_user.id,
        photo_url=photo_url,
    )
    db.add(student)
    db.flush()

    # Crear exam_results pendientes para exámenes ya existentes en el curso
    exams = db.query(Exam).filter(
        Exam.course_id == course_id,
        Exam.teacher_id == current_user.id,
    ).all()
    for exam in exams:
        exists = db.query(ExamResult).filter(
            ExamResult.exam_id == exam.id,
            ExamResult.student_id == student.id,
        ).first()
        if not exists:
            db.add(ExamResult(
                exam_id=exam.id, student_id=student.id,
                status=ExamStatusEnum.pending, image_urls=json.dumps([]),
            ))

    db.commit()
    db.refresh(student)
    return _student_to_dict(student)


@router.patch("/estudiantes/{student_id}")
async def update_student(
    student_id: str,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_docente),
):
    s = db.query(Student).filter(
        Student.id == student_id,
        Student.institution_id == current_user.institution_id,
    ).first()
    if not s:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")

    body = await request.json()
    name = body.get("fullName") or body.get("full_name")
    photo_url = body.get("photoUrl") or body.get("photo_url")

    if name:
        s.full_name = name.strip()
    if "photoUrl" in body or "photo_url" in body:
        s.photo_url = photo_url

    db.commit()
    db.refresh(s)
    return _student_to_dict(s)


# ─── EXÁMENES ─────────────────────────────────────────────────────────────────
@router.post("/examenes", status_code=201)
async def create_exam(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_docente),
):
    """
    Crea el examen Y genera automáticamente un exam_result
    pendiente para CADA estudiante del curso.
    """
    inst_id = _teacher_institution(current_user)
    body = await request.json()
    name = (body.get("name") or "").strip()
    course_id = body.get("courseId") or body.get("course_id")

    if not name:
        raise HTTPException(status_code=400, detail="El nombre del examen es requerido")
    if not course_id:
        raise HTTPException(status_code=400, detail="Se requiere courseId")

    exam = Exam(
        name=name,
        course_id=course_id,
        teacher_id=current_user.id,
        institution_id=inst_id,
    )
    db.add(exam)
    db.flush()  # Obtener ID sin commit aún

    # Crear un cajón pendiente por cada estudiante del curso
    students = db.query(Student).filter(
        Student.course_id == course_id,
        Student.institution_id == inst_id,
    ).all()

    for student in students:
        db.add(ExamResult(
            exam_id=exam.id,
            student_id=student.id,
            status=ExamStatusEnum.pending,
            image_urls=json.dumps([]),
        ))

    db.commit()
    db.refresh(exam)
    return {
        "id": exam.id,
        "name": exam.name,
        "courseId": exam.course_id,
        "teacherId": exam.teacher_id,
        "createdAt": exam.created_at.isoformat(),
        "studentsCount": len(students),
    }

@router.delete("/examenes/{exam_id}", status_code=204)
def delete_exam(
    exam_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_docente),
):
    """Elimina un examen y todos sus resultados."""
    exam = db.query(Exam).filter(
        Exam.id == exam_id,
        Exam.teacher_id == current_user.id,
    ).first()
    if not exam:
        raise HTTPException(status_code=404, detail="Examen no encontrado")
    # Eliminar resultados primero
    db.query(ExamResult).filter(ExamResult.exam_id == exam_id).delete()
    db.delete(exam)
    db.commit()


@router.patch("/examenes/{exam_id}")
async def update_exam(
    exam_id: str,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_docente),
):
    """Edita el nombre de un examen."""
    exam = db.query(Exam).filter(
        Exam.id == exam_id,
        Exam.teacher_id == current_user.id,
    ).first()
    if not exam:
        raise HTTPException(status_code=404, detail="Examen no encontrado")
    body = await request.json()
    name = (body.get("name") or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="El nombre es requerido")
    exam.name = name
    db.commit()
    db.refresh(exam)
    return {"id": exam.id, "name": exam.name, "courseId": exam.course_id}


@router.delete("/examenes/resultados/{result_id}/contenido", status_code=204)
def clear_exam_result(
    result_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_docente),
):
    """Limpia el contenido de un resultado para volver a escanear."""
    result = db.query(ExamResult).filter(ExamResult.id == result_id).first()
    if not result:
        raise HTTPException(status_code=404, detail="Resultado no encontrado")
    result.ocr_raw_text = None
    result.problems_json = "[]"
    result.final_score = None
    result.grade_color = None
    result.teacher_notes = None
    result.image_urls = "[]"
    result.status = ExamStatusEnum.pending
    db.commit()


@router.get("/cursos/{course_id}/examenes")
def list_course_exams(
    course_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_docente),
):
    """Lista todos los exámenes de un curso."""
    exams = db.query(Exam).filter(
        Exam.course_id == course_id,
        Exam.teacher_id == current_user.id,
    ).order_by(Exam.created_at.desc()).all()
    return [
        {
            "id": e.id,
            "name": e.name,
            "courseId": e.course_id,
            "createdAt": e.created_at.isoformat(),
        }
        for e in exams
    ]


@router.get("/estudiantes/{student_id}/examenes")
def list_student_exams(
    student_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_docente),
):
    """
    Devuelve todos los cajones de examen de un estudiante,
    incluyendo el nombre del examen.
    """
    results = db.query(ExamResult).filter(
        ExamResult.student_id == student_id,
    ).order_by(ExamResult.created_at.desc()).all()

    output = []
    for r in results:
        exam = db.query(Exam).filter(Exam.id == r.exam_id).first()
        exam_name = exam.name if exam else "Examen"
        output.append(_result_to_dict(r, exam_name))
    return output


# ─── ESCANEAR (OCR) ───────────────────────────────────────────────────────────
@router.post("/examenes/escanear")
async def scan_exam(
    request: Request,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_docente),
):
    """
    Recibe imágenes base64, actualiza el exam_result existente a 'processing'
    y lanza el OCR en background. Devuelve el resultId inmediatamente.
    """
    inst_id = _teacher_institution(current_user)
    body = await request.json()
    exam_result_id = body.get("examResultId") or body.get("exam_result_id")
    exam_id = body.get("examId") or body.get("exam_id")
    student_id = body.get("studentId") or body.get("student_id")
    course_id = body.get("courseId") or body.get("course_id")
    images: list = body.get("images", [])

    if not images:
        raise HTTPException(status_code=400, detail="Se requiere al menos una imagen")

    result = None

    # 1. Buscar result por ID directo
    if exam_result_id:
        result = db.query(ExamResult).filter(ExamResult.id == exam_result_id).first()

    # 2. Buscar result por exam+student
    if not result and exam_id and student_id:
        result = db.query(ExamResult).filter(
            ExamResult.exam_id == exam_id,
            ExamResult.student_id == student_id,
        ).first()

    # 3. Crear result si no existe (fallback)
    if not result:
        if not exam_id:
            # Crear examen genérico
            new_exam = Exam(
                name="Examen sin título",
                course_id=course_id or "",
                teacher_id=current_user.id,
                institution_id=inst_id,
            )
            db.add(new_exam)
            db.flush()
            exam_id = new_exam.id

        result = ExamResult(
            exam_id=exam_id,
            student_id=student_id,
            status=ExamStatusEnum.processing,
            image_urls=json.dumps([]),
        )
        db.add(result)
        db.commit()
        db.refresh(result)
    else:
        result.status = ExamStatusEnum.processing
        db.commit()

    result_id = result.id
    background_tasks.add_task(
        _process_ocr_background,
        result_id=result_id,
        images=images,
    )

    return {"resultId": result_id, "status": "processing"}


def _process_ocr_background(result_id: str, images: list):
    """
    Ejecuta el OCR y guarda el resultado completo.
    Equivalente al bloque try/catch del endpoint /scan-exam del backend viejo.
    """
    from ..database import SessionLocal
    db = SessionLocal()
    result = None
    try:
        result = db.query(ExamResult).filter(ExamResult.id == result_id).first()
        if not result:
            return

        # 1. OCR
        ocr_text = extract_text_from_images(images)
        result.ocr_raw_text = ocr_text

        # 2. Parsear problemas (parseExamSheet del backend viejo)
        problems = detect_math_problems(ocr_text)
        result.problems_json = json.dumps(problems, ensure_ascii=False)

        # 3. Calificar
        score = calculate_final_score(problems)
        result.final_score = score
        result.grade_color = "green" if score >= PASSING_GRADE else "red"

        # 4. Plan pedagógico (generateTeachingPlanLogic del backend viejo)
        plan = generate_teaching_plan(problems)

        # 5. Guardar plan en teacher_notes con separador
        result.teacher_notes = f"---PLAN---\n{plan}"

        # 6. Guardar URLs de imágenes (marcadores — en prod serían paths reales)
        result.image_urls = json.dumps([
            f"result_{result_id}_img{i}" for i in range(len(images))
        ])

        result.status = ExamStatusEnum.graded
        db.commit()

    except Exception as e:
        if result:
            result.status = ExamStatusEnum.reviewed
            result.ocr_raw_text = (result.ocr_raw_text or "") + f"\n[Error OCR: {e}]"
            db.commit()
    finally:
        db.close()


# ─── VER RESULTADO ────────────────────────────────────────────────────────────
@router.get("/examenes/resultados/{result_id}")
def get_exam_result(
    result_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_docente),
):
    result = db.query(ExamResult).filter(ExamResult.id == result_id).first()
    if not result:
        raise HTTPException(status_code=404, detail="Resultado no encontrado")
    exam = db.query(Exam).filter(Exam.id == result.exam_id).first()
    return _result_to_dict(result, exam.name if exam else "")


# ─── GUARDAR REVISIÓN DEL DOCENTE ─────────────────────────────────────────────
@router.patch("/examenes/resultados/{result_id}")
async def update_exam_result(
    result_id: str,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_docente),
):
    result = db.query(ExamResult).filter(ExamResult.id == result_id).first()
    if not result:
        raise HTTPException(status_code=404, detail="Resultado no encontrado")

    body = await request.json()
    problems = body.get("problems")
    final_score = body.get("finalScore") or body.get("final_score")
    teacher_notes = (body.get("teacherNotes") or body.get("teacher_notes") or "").strip()

    if problems is not None:
        result.problems_json = json.dumps(problems, ensure_ascii=False)
        # Recalcular nota con las correcciones del docente
        if final_score is None:
            final_score = calculate_final_score(problems)

    if final_score is not None:
        result.final_score = float(final_score)
        result.grade_color = "green" if float(final_score) >= PASSING_GRADE else "red"

    # Regenerar plan pedagógico con problemas actualizados
    probs = problems or json.loads(result.problems_json or "[]")
    new_plan = generate_teaching_plan(probs)

    # Guardar notas del docente + plan con separador
    result.teacher_notes = f"{teacher_notes}---PLAN---\n{new_plan}"
    result.status = ExamStatusEnum.graded

    db.commit()
    db.refresh(result)
    exam = db.query(Exam).filter(Exam.id == result.exam_id).first()
    return _result_to_dict(result, exam.name if exam else "")


# ─── REPORTES ─────────────────────────────────────────────────────────────────
@router.get("/reportes")
def get_reports(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_docente),
):
    inst_id = _teacher_institution(current_user)
    courses = db.query(Course).filter(
        Course.institution_id == inst_id,
        Course.teacher_id == current_user.id,
    ).all()
    course_ids = [c.id for c in courses]

    exam_ids_q = db.query(Exam.id).filter(Exam.teacher_id == current_user.id)
    all_results = db.query(ExamResult).filter(
        ExamResult.exam_id.in_(exam_ids_q),
        ExamResult.final_score.isnot(None),
    ).all()

    scores = [r.final_score for r in all_results if r.final_score]
    aprobados = sum(1 for s in scores if s >= PASSING_GRADE)
    reprobados = len(scores) - aprobados
    promedio = round(sum(scores) / len(scores), 2) if scores else 0.0

    rangos = [("1.0-1.9", 1.0, 2.0), ("2.0-2.9", 2.0, 3.0), ("3.0-3.9", 3.0, 4.0), ("4.0-5.0", 4.0, 5.1)]
    distribucion = [
        {"rango": r, "cantidad": sum(1 for s in scores if lo <= s < hi)}
        for r, lo, hi in rangos
    ]

    promedios_por_curso = []
    for course in courses:
        c_ids = [
            e.id for e in db.query(Exam).filter(
                Exam.course_id == course.id,
                Exam.teacher_id == current_user.id,
            ).all()
        ]
        c_scores = [
            r.final_score for r in db.query(ExamResult.final_score).filter(
                ExamResult.exam_id.in_(c_ids),
                ExamResult.final_score.isnot(None),
            ).all() if r.final_score
        ]
        if c_scores:
            promedios_por_curso.append({
                "courseId": course.id, "courseName": course.name,
                "promedio": round(sum(c_scores) / len(c_scores), 2),
                "count": len(c_scores),
            })

    total_students = (
        db.query(Student).filter(Student.course_id.in_(course_ids)).count()
        if course_ids else 0
    )

    return {
        "totalExamenes": len(all_results),
        "totalEstudiantes": total_students,
        "aprobados": aprobados,
        "reprobados": reprobados,
        "promedioGeneral": promedio,
        "promediosPorCurso": promedios_por_curso,
        "distribucion": distribucion,
    }

@router.get("/cursos/{course_id}/exportar-notas")
def export_course_grades(
    course_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_docente),
):
    """Devuelve JSON con datos de notas para preview y exportación."""
    inst_id = _teacher_institution(current_user)
    course = db.query(Course).filter(Course.id == course_id, Course.institution_id == inst_id).first()
    if not course:
        raise HTTPException(status_code=404, detail="Curso no encontrado")

    students = db.query(Student).filter(Student.course_id == course_id).order_by(Student.full_name).all()
    exams = db.query(Exam).filter(Exam.course_id == course_id, Exam.teacher_id == current_user.id).order_by(Exam.created_at).all()

    rows = []
    for student in students:
        scores = []
        exam_data = []
        for exam in exams:
            result = db.query(ExamResult).filter(
                ExamResult.exam_id == exam.id, ExamResult.student_id == student.id
            ).first()
            score = result.final_score if result and result.final_score is not None else None
            exam_data.append({"examName": exam.name, "score": score, "status": result.status.value if result else "pending"})
            if score is not None:
                scores.append(score)
        rows.append({
            "studentName": student.full_name,
            "exams": exam_data,
            "average": round(sum(scores) / len(scores), 1) if scores else None,
        })

    return {
        "courseName": course.name, "grade": course.grade, "group": course.group,
        "exams": [e.name for e in exams], "rows": rows,
        "exportedAt": date.today().isoformat(),
    }


@router.get("/cursos/{course_id}/exportar-xlsx")
def export_course_xlsx(
    course_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_docente),
):
    """Genera Excel y lo sirve como archivo descargable."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    inst_id = _teacher_institution(current_user)
    course = db.query(Course).filter(
        Course.id == course_id, Course.institution_id == inst_id
    ).first()
    if not course:
        raise HTTPException(status_code=404, detail="Curso no encontrado")

    students = db.query(Student).filter(
        Student.course_id == course_id
    ).order_by(Student.full_name).all()

    exams = db.query(Exam).filter(
        Exam.course_id == course_id,
        Exam.teacher_id == current_user.id,
    ).order_by(Exam.created_at).all()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Curso {course.grade}-{course.group}"

    BLUE     = "1A3C5E"
    BLUE2    = "2E6DA4"
    GREEN_BG = "C6EFCE"; GREEN_FG = "276221"
    RED_BG   = "FFC7CE"; RED_FG   = "9C0006"
    PEND_BG  = "FFEB9C"; PEND_FG  = "9C5700"
    ALT      = "F2F6FA"
    thin     = Side(style="thin", color="BBBBBB")
    brd      = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft      = Alignment(horizontal="left",   vertical="center")

    total_cols = 1 + len(exams) + 1  # nombre + examenes + promedio

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws.cell(row=1, column=1, value=f"EMAI-APP — Notas Curso {course.grade}-{course.group}")
    c.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=BLUE)
    c.alignment = ctr
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    c2 = ws.cell(row=2, column=1, value=f"Exportado: {date.today().strftime('%d/%m/%Y')}")
    c2.font = Font(name="Arial", size=10, color="FFFFFF")
    c2.fill = PatternFill("solid", fgColor=BLUE)
    c2.alignment = ctr
    ws.row_dimensions[2].height = 16

    # Cabeceras
    headers = ["Estudiante"] + [e.name for e in exams] + ["Promedio"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", fgColor=BLUE2)
        cell.alignment = ctr if col > 1 else lft
        cell.border    = brd
    ws.row_dimensions[3].height = 22

    # Datos
    for ri, student in enumerate(students):
        row = 4 + ri
        fill_row = PatternFill("solid", fgColor=ALT) if ri % 2 else None

        # Nombre
        nc = ws.cell(row=row, column=1, value=student.full_name)
        nc.font = Font(name="Arial", size=11)
        nc.alignment = lft
        nc.border = brd
        if fill_row: nc.fill = fill_row

        scores = []
        for ci, exam in enumerate(exams):
            result = db.query(ExamResult).filter(
                ExamResult.exam_id == exam.id,
                ExamResult.student_id == student.id,
            ).first()
            score = result.final_score if result and result.final_score is not None else None
            col = ci + 2

            if score is not None:
                val = float(score)
                scores.append(val)
                passed = val >= 3.0
                sc = ws.cell(row=row, column=col, value=val)
                sc.number_format = "0.0"
                sc.font = Font(name="Arial", bold=True, size=11,
                               color=GREEN_FG if passed else RED_FG)
                sc.fill = PatternFill("solid", fgColor=GREEN_BG if passed else RED_BG)
            else:
                sc = ws.cell(row=row, column=col, value="Pendiente")
                sc.font = Font(name="Arial", size=10, color=PEND_FG, italic=True)
                sc.fill = PatternFill("solid", fgColor=PEND_BG)
            sc.alignment = ctr
            sc.border = brd

        # Promedio
        avg_col = len(exams) + 2
        if scores:
            avg = round(sum(scores) / len(scores), 1)
            passed = avg >= 3.0
            ac = ws.cell(row=row, column=avg_col, value=avg)
            ac.number_format = "0.0"
            ac.font = Font(name="Arial", bold=True, size=12,
                           color=GREEN_FG if passed else RED_FG)
            ac.fill = PatternFill("solid", fgColor=GREEN_BG if passed else RED_BG)
        else:
            ac = ws.cell(row=row, column=avg_col, value="—")
            ac.font = Font(name="Arial", size=11, color="888888")
        ac.alignment = ctr
        ac.border = brd
        ws.row_dimensions[row].height = 20

    # Anchos
    ws.column_dimensions["A"].width = 28
    for ci in range(len(exams)):
        col_letter = get_column_letter(ci + 2)
        ws.column_dimensions[col_letter].width = max(12, min(len(exams[ci].name) + 2, 24))
    ws.column_dimensions[get_column_letter(len(exams) + 2)].width = 12
    ws.freeze_panes = f"B4"

    # Serializar
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"notas_{course.grade}-{course.group}_{date.today().isoformat()}.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

# ─── SOPORTE ──────────────────────────────────────────────────────────────────
@router.post("/soporte", status_code=201)
async def send_support(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_docente),
):
    body = await request.json()
    message = (body.get("message") or "").strip()
    if not message:
        raise HTTPException(status_code=400, detail="El mensaje es requerido")
    msg = SupportMessage(
        from_user_id=current_user.id,
        from_username=current_user.username,
        from_role=current_user.role.value,
        institution_id=current_user.institution_id,
        message=message,
    )
    db.add(msg)
    db.commit()
    return {"ok": True}
